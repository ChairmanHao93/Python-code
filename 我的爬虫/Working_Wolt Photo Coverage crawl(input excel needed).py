import os
import requests
from bs4 import BeautifulSoup
import openpyxl

# Function to process a single URL and calculate product photo coverage
def calculate_photo_coverage(url):
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        div_elements = soup.find_all("div", class_="sc-c89e7f2c-0 gDJLOI")
        
        count_with_images = 0
        for div in div_elements:
            if div.find("img"):
                count_with_images += 1
        
        total_div_elements = len(div_elements)
        percentage_photo_coverage = (count_with_images / total_div_elements) * 100
        
        return count_with_images, total_div_elements, percentage_photo_coverage
    
    else:
        return None, None, None

# Path to the Excel file
excel_file_path = r"C:\Users\h.zhang.2\Desktop\Wolt pages.xlsx"

# Load the Excel file using openpyxl
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Create lists to store results
urls = []
product_with_photo = []
product_listed = []
percentage_photo_covered = []

# Iterate through rows and read URLs from the "Wolt page url" column
for row in sheet.iter_rows(min_row=2, values_only=True, max_col=1):
    url = row[0]
    if url:
        urls.append(url)
        count_with_images, total_div_elements, percentage_photo_coverage = calculate_photo_coverage(url)
        
        if count_with_images is not None:
            product_with_photo.append(count_with_images)
            product_listed.append(total_div_elements)
            percentage_photo_covered.append(percentage_photo_coverage)
        else:
            product_with_photo.append("N/A")
            product_listed.append("N/A")
            percentage_photo_covered.append("N/A")

# Create a new Excel workbook and sheet to store results
result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active

# Write headers
result_sheet.cell(row=1, column=1, value="Wolt page url")
result_sheet.cell(row=1, column=2, value="Product with photo")
result_sheet.cell(row=1, column=3, value="Product listed")
result_sheet.cell(row=1, column=4, value="Percentage of photo covered product")

# Write data
for i, url in enumerate(urls):
    result_sheet.cell(row=i + 2, column=1, value=url)
    result_sheet.cell(row=i + 2, column=2, value=product_with_photo[i])
    result_sheet.cell(row=i + 2, column=3, value=product_listed[i])
    result_sheet.cell(row=i + 2, column=4, value=percentage_photo_covered[i])

# Save the result Excel file
result_excel_path = os.path.join(os.path.dirname(excel_file_path), "Wolt_product_photo_coverage.xlsx")
result_workbook.save(result_excel_path)
print(f"Results saved in Excel file: {result_excel_path}")
