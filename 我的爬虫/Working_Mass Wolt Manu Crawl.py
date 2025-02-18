import os
import requests
from bs4 import BeautifulSoup
import openpyxl

# Load URLs from the Excel file
excel_file_path = r"C:\Users\h.zhang.2\Desktop\Wolt pages.xlsx"
wb_urls = openpyxl.load_workbook(excel_file_path)
ws_urls = wb_urls.active

# Create the output folder if it doesn't exist
output_folder = r"C:\Users\h.zhang.2\Desktop\All Wolt Manu"
os.makedirs(output_folder, exist_ok=True)

# Iterate through URLs and scrape data
for row in ws_urls.iter_rows(min_row=2, values_only=True):
    url = row[0]
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all h3 and span elements using specified data attributes
    h3_elements = soup.find_all('h3', {'data-test-id': 'horizontal-item-card-header'})
    span_elements = soup.find_all('span', {'data-test-id': 'horizontal-item-card-price'})

    # Find all p elements with the new class
    p_elements = soup.find_all('p', class_='sc-866b5984-0 bwtZvA')

    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the headers
    ws.cell(row=1, column=1, value="Dish name")
    ws.cell(row=1, column=2, value="Description")
    ws.cell(row=1, column=3, value="Price")

    # Write the data to the Excel sheet
    for idx, (h3, p, span) in enumerate(zip(h3_elements, p_elements, span_elements), start=2):
        # Sanitize dish name to remove problematic characters
        dish_name = h3.text.strip()
        dish_name = ''.join([c for c in dish_name if c.isprintable()])

        ws.cell(row=idx, column=1, value=dish_name)
        ws.cell(row=idx, column=2, value=p.text.strip())
        ws.cell(row=idx, column=3, value=span.text.strip())

 # Save the Excel file
url_parts = url.split("/")
sheet_name = url_parts[-1] if url_parts[-1] else url_parts[-2]  # Use the last part of the URL after the last "/" as the sheet name
excel_filename = os.path.join(output_folder, f"{sheet_name}.xlsx")
wb.save(excel_filename)
print(f"Data saved to '{excel_filename}'")

print("All data saved.")
