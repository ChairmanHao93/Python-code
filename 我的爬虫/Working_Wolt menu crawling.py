import requests
from bs4 import BeautifulSoup
import openpyxl

# Define the URL of the website
url = "https://wolt.com/fi/fin/lahti/restaurant/hua-du-lahti"
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')

# Find all <button> elements with the specified data-test-id attribute
button_elements = soup.find_all('button', {'data-test-id': 'horizontal-item-card-button'})

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active

# Write the headers
ws.cell(row=1, column=1, value="Dish name")
ws.cell(row=1, column=2, value="Description")
ws.cell(row=1, column=3, value="Price")

# Write the data to the Excel sheet
for idx, button in enumerate(button_elements, start=2):
    # Find elements within the <button> element
    h3 = button.find('h3', {'data-test-id': 'horizontal-item-card-header'})
    
    # Find <p> element with class 'sc-96436756-2 jnnBXN' 
    p = button.find('p', class_='sc-96436756-2 jnnBXN') 
    span = button.find('span', {'data-test-id': 'horizontal-item-card-price'})
    
    ws.cell(row=idx, column=1, value=h3.text.strip() if h3 else "")
    ws.cell(row=idx, column=2, value=p.text.strip() if p else "")
    ws.cell(row=idx, column=3, value=span.text.strip() if span else "")

# Save the Excel file
excel_filename = "MenuCB.xlsx"
wb.save(excel_filename)
print(f"Data saved to '{excel_filename}'")
