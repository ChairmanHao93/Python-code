import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl

# Define the output folder path
output_folder = r"C:\Users\h.zhang.2\Desktop\Foodora manu"

# Create a WebDriver instance (ChromeDriver in this case)
driver = webdriver.Chrome()  # Make sure you have ChromeDriver installed

# Navigate to the URL
url = "https://wolt.com/fi/fin/vantaa/restaurant/konnichiwa-tikkurila-uusi"
driver.get(url)

# Handle the popup (if it appears)
try:
    popup_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "button.bds-c-modal__close-button"))
    )
    popup_button.click()
except:
    print("Popup button not found or not interactable")

# Wait a bit to let the page load
driver.implicitly_wait(5)

# Create a BeautifulSoup instance from the page source
soup = BeautifulSoup(driver.page_source, 'html.parser')

# Find all elements for product names, descriptions, and prices
product_name_elements = soup.find_all('span', {'data-testid': 'menu-product-name'})
description_elements = soup.find_all('p', {'data-testid': 'menu-product-description'})
price_elements = soup.find_all('p', {'data-testid': 'menu-product-price'})


# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active

# Write the headers
ws.cell(row=1, column=1, value="Product Name")
ws.cell(row=1, column=2, value="Description")
ws.cell(row=1, column=3, value="Price")

# Write the data to the Excel sheet
for idx, (product_name, description, price) in enumerate(zip(product_name_elements, description_elements, price_elements), start=2):
    # Debugging: Print the content of elements
    print(f"Product name: {product_name.text.strip()}")
    print(f"Description: {description.text.strip()}")
    print(f"Price: {price.text.strip()}")

    ws.cell(row=idx, column=1, value=product_name.text.strip())
    ws.cell(row=idx, column=2, value=description.text.strip())
    ws.cell(row=idx, column=3, value=price.text.strip())

# Save the Excel file
url_parts = url.split("/")
sheet_name = url_parts[-1] if url_parts[-1] else url_parts[-2]  # Use the last part of the URL after the last "/" as the sheet name
excel_filename = os.path.join(output_folder, f"{sheet_name}.xlsx")
wb.save(excel_filename)
print(f"Data saved to '{excel_filename}'")

# Quit the WebDriver
driver.quit()
